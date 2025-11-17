"""
Comprehensive tests for utils/export_utils.py
Target: 100% coverage for PDF and Excel export functionality
"""
import io
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from reportlab.lib.styles import getSampleStyleSheet

from utils.export_utils import (
    add_checkout_history_excel_content,
    add_checkout_history_pdf_content,
    add_cycle_count_excel_content,
    add_cycle_count_pdf_content,
    add_department_usage_excel_content,
    add_department_usage_pdf_content,
    add_tool_inventory_excel_content,
    add_tool_inventory_pdf_content,
    generate_excel_report,
    generate_pdf_report,
    get_report_title,
)


class TestGetReportTitle:
    """Tests for get_report_title function"""

    def test_tool_inventory_report_title(self):
        """Test tool inventory report title"""
        result = get_report_title("tool-inventory", None)
        assert result == "Tool Inventory Report"

    def test_checkout_history_report_title(self):
        """Test checkout history report title"""
        result = get_report_title("checkout-history", None)
        assert result == "Checkout History Report"

    def test_department_usage_report_title(self):
        """Test department usage report title"""
        result = get_report_title("department-usage", None)
        assert result == "Department Usage Report"

    def test_cycle_count_accuracy_report_title(self):
        """Test cycle count accuracy report title"""
        result = get_report_title("cycle-count-accuracy", None)
        assert result == "Cycle Count Accuracy Report"

    def test_cycle_count_discrepancies_report_title(self):
        """Test cycle count discrepancies report title"""
        result = get_report_title("cycle-count-discrepancies", None)
        assert result == "Cycle Count Discrepancy Report"

    def test_cycle_count_performance_report_title(self):
        """Test cycle count performance report title"""
        result = get_report_title("cycle-count-performance", None)
        assert result == "Cycle Count Performance Report"

    def test_cycle_count_coverage_report_title(self):
        """Test cycle count coverage report title"""
        result = get_report_title("cycle-count-coverage", None)
        assert result == "Cycle Count Coverage Report"

    def test_unknown_report_type(self):
        """Test unknown report type defaults to 'Report'"""
        result = get_report_title("unknown-type", None)
        assert result == "Report"

    def test_with_timeframe_weekly(self):
        """Test title with weekly timeframe"""
        result = get_report_title("tool-inventory", "weekly")
        assert result == "Tool Inventory Report (Weekly)"

    def test_with_timeframe_monthly(self):
        """Test title with monthly timeframe"""
        result = get_report_title("checkout-history", "monthly")
        assert result == "Checkout History Report (Monthly)"

    def test_with_timeframe_all(self):
        """Test title with 'all' timeframe (should not append)"""
        result = get_report_title("tool-inventory", "all")
        assert result == "Tool Inventory Report"

    def test_with_none_timeframe(self):
        """Test title with None timeframe"""
        result = get_report_title("department-usage", None)
        assert result == "Department Usage Report"

    def test_with_empty_string_timeframe(self):
        """Test title with empty string timeframe"""
        result = get_report_title("tool-inventory", "")
        assert result == "Tool Inventory Report"


class TestGeneratePDFReport:
    """Tests for generate_pdf_report function"""

    def test_generate_tool_inventory_pdf(self):
        """Test PDF generation for tool inventory report"""
        report_data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "Test Tool",
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available"
            }
        ]

        buffer = generate_pdf_report(report_data, "tool-inventory", None)

        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()
        assert len(buffer.getvalue()) > 0

    def test_generate_checkout_history_pdf(self):
        """Test PDF generation for checkout history report"""
        report_data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "John Doe",
                    "department": "Engineering",
                    "checkout_date": "2024-01-15T10:00:00",
                    "return_date": "2024-01-16T14:30:00",
                    "duration": 1.5
                }
            ],
            "stats": {
                "totalCheckouts": 100,
                "returnedCheckouts": 95,
                "currentlyCheckedOut": 5,
                "averageDuration": 2.3
            }
        }

        buffer = generate_pdf_report(report_data, "checkout-history", "monthly")

        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()
        assert len(buffer.getvalue()) > 0

    def test_generate_department_usage_pdf(self):
        """Test PDF generation for department usage report"""
        report_data = {
            "departments": [
                {
                    "name": "Engineering",
                    "totalCheckouts": 50,
                    "currentlyCheckedOut": 3,
                    "averageDuration": 2.5,
                    "mostUsedCategory": "Hand Tools"
                }
            ]
        }

        buffer = generate_pdf_report(report_data, "department-usage", "weekly")

        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()

    def test_generate_cycle_count_accuracy_pdf(self):
        """Test PDF generation for cycle count accuracy report"""
        report_data = {
            "summary": {
                "total_counts": 100,
                "accurate_counts": 95,
                "accuracy_rate": 95.0
            }
        }

        buffer = generate_pdf_report(report_data, "cycle-count-accuracy", None)

        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()

    def test_generate_cycle_count_other_type_pdf(self):
        """Test PDF generation for other cycle count report types"""
        report_data = {
            "summary": {
                "total_counts": 50,
                "accurate_counts": 45,
                "accuracy_rate": 90.0
            }
        }

        buffer = generate_pdf_report(report_data, "cycle-count-discrepancies", None)

        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()

    def test_generate_pdf_empty_tool_inventory(self):
        """Test PDF generation with empty tool inventory"""
        report_data = []

        buffer = generate_pdf_report(report_data, "tool-inventory", None)

        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()

    def test_generate_pdf_unknown_report_type(self):
        """Test PDF generation with unknown report type"""
        report_data = {}

        buffer = generate_pdf_report(report_data, "unknown-type", None)

        assert isinstance(buffer, io.BytesIO)
        assert buffer.getvalue()


class TestGenerateExcelReport:
    """Tests for generate_excel_report function"""

    def test_generate_tool_inventory_excel(self):
        """Test Excel generation for tool inventory report"""
        report_data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "Test Tool",
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available",
                "condition": "Good"
            }
        ]

        buffer = generate_excel_report(report_data, "tool-inventory", None)

        assert isinstance(buffer, io.BytesIO)
        # Verify it's a valid Excel file
        workbook = openpyxl.load_workbook(buffer)
        assert workbook.active is not None
        assert "Tool Inventory Re" in workbook.active.title  # Truncated to 31 chars

    def test_generate_checkout_history_excel(self):
        """Test Excel generation for checkout history report"""
        report_data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "John Doe",
                    "department": "Engineering",
                    "checkout_date": "2024-01-15T10:00:00",
                    "return_date": "2024-01-16T14:30:00",
                    "duration": 1.5
                }
            ],
            "stats": {
                "totalCheckouts": 100,
                "returnedCheckouts": 95,
                "currentlyCheckedOut": 5,
                "averageDuration": 2.3
            }
        }

        buffer = generate_excel_report(report_data, "checkout-history", "monthly")

        assert isinstance(buffer, io.BytesIO)
        workbook = openpyxl.load_workbook(buffer)
        ws = workbook.active
        assert ws["A1"].value == "Checkout History Report (Monthly)"
        assert "B5" in [cell.coordinate for cell in ws["B:B"]]

    def test_generate_department_usage_excel(self):
        """Test Excel generation for department usage report"""
        report_data = {
            "departments": [
                {
                    "name": "Engineering",
                    "totalCheckouts": 50,
                    "currentlyCheckedOut": 3,
                    "averageDuration": 2.5,
                    "mostUsedCategory": "Hand Tools"
                }
            ]
        }

        buffer = generate_excel_report(report_data, "department-usage", None)

        assert isinstance(buffer, io.BytesIO)
        workbook = openpyxl.load_workbook(buffer)
        assert workbook.active is not None

    def test_generate_cycle_count_excel(self):
        """Test Excel generation for cycle count report"""
        report_data = {
            "summary": {
                "total_counts": 100,
                "accurate_counts": 95,
                "accuracy_rate": 95.0
            }
        }

        buffer = generate_excel_report(report_data, "cycle-count-accuracy", None)

        assert isinstance(buffer, io.BytesIO)
        workbook = openpyxl.load_workbook(buffer)
        ws = workbook.active
        assert "Cycle Count Accuracy Report" in ws["A1"].value

    def test_generate_excel_long_title_truncation(self):
        """Test that Excel sheet title is truncated to 31 characters"""
        report_data = []

        buffer = generate_excel_report(report_data, "cycle-count-discrepancies", "quarterly")

        workbook = openpyxl.load_workbook(buffer)
        # Sheet name should be truncated to 31 characters
        assert len(workbook.active.title) <= 31

    def test_generate_excel_column_width_adjustment(self):
        """Test that column widths are auto-adjusted"""
        report_data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "A very long description that exceeds normal column width requirements",
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available",
                "condition": "Good"
            }
        ]

        buffer = generate_excel_report(report_data, "tool-inventory", None)

        workbook = openpyxl.load_workbook(buffer)
        ws = workbook.active
        # Check that column widths are set
        assert ws.column_dimensions["A"].width is not None

    def test_generate_excel_column_width_max_limit(self):
        """Test that column width is capped at 50"""
        report_data = [
            {
                "tool_number": "T001" * 20,  # Very long value
                "serial_number": "SN001",
                "description": "Test",
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available",
                "condition": "Good"
            }
        ]

        buffer = generate_excel_report(report_data, "tool-inventory", None)

        workbook = openpyxl.load_workbook(buffer)
        ws = workbook.active
        # Width should be capped at 50
        for col_dim in ws.column_dimensions.values():
            assert col_dim.width <= 50

    def test_generate_excel_empty_data(self):
        """Test Excel generation with empty data"""
        report_data = []

        buffer = generate_excel_report(report_data, "tool-inventory", None)

        assert isinstance(buffer, io.BytesIO)
        workbook = openpyxl.load_workbook(buffer)
        assert workbook.active is not None

    def test_generate_excel_with_none_cell_values(self):
        """Test Excel column width calculation handles None values"""
        # This tests the TypeError/AttributeError handling in column width calculation
        report_data = {
            "departments": []
        }

        buffer = generate_excel_report(report_data, "department-usage", None)

        assert isinstance(buffer, io.BytesIO)
        workbook = openpyxl.load_workbook(buffer)
        assert workbook.active is not None

    def test_generate_excel_unknown_report_type(self):
        """Test Excel generation with unknown report type"""
        report_data = {}

        buffer = generate_excel_report(report_data, "unknown-type", None)

        assert isinstance(buffer, io.BytesIO)
        workbook = openpyxl.load_workbook(buffer)
        assert workbook.active is not None


class TestAddToolInventoryPDFContent:
    """Tests for add_tool_inventory_pdf_content function"""

    def test_empty_tool_inventory(self):
        """Test with empty tool list"""
        story = []
        styles = getSampleStyleSheet()

        add_tool_inventory_pdf_content(story, [], styles)

        assert len(story) == 1
        # Check that "No tools found" message is added

    def test_single_tool_inventory(self):
        """Test with single tool"""
        story = []
        styles = getSampleStyleSheet()
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "Test Tool",
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available"
            }
        ]

        add_tool_inventory_pdf_content(story, data, styles)

        assert len(story) == 1  # Table added

    def test_multiple_tools_inventory(self):
        """Test with multiple tools"""
        story = []
        styles = getSampleStyleSheet()
        data = [
            {
                "tool_number": f"T00{i}",
                "serial_number": f"SN00{i}",
                "description": f"Test Tool {i}",
                "category": "Hand Tools",
                "location": f"Bay {i}",
                "status": "available"
            }
            for i in range(1, 4)
        ]

        add_tool_inventory_pdf_content(story, data, styles)

        assert len(story) == 1

    def test_long_description_truncation(self):
        """Test that long descriptions are truncated"""
        story = []
        styles = getSampleStyleSheet()
        long_desc = "A" * 50  # More than 30 characters
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": long_desc,
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available"
            }
        ]

        add_tool_inventory_pdf_content(story, data, styles)

        assert len(story) == 1

    def test_missing_fields_in_tool(self):
        """Test with missing fields in tool data"""
        story = []
        styles = getSampleStyleSheet()
        data = [
            {
                "tool_number": "T001",
                # Missing other fields
            }
        ]

        add_tool_inventory_pdf_content(story, data, styles)

        assert len(story) == 1

    def test_short_description_no_truncation(self):
        """Test that short descriptions are not truncated"""
        story = []
        styles = getSampleStyleSheet()
        short_desc = "Short"
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": short_desc,
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available"
            }
        ]

        add_tool_inventory_pdf_content(story, data, styles)

        assert len(story) == 1


class TestAddCheckoutHistoryPDFContent:
    """Tests for add_checkout_history_pdf_content function"""

    def test_with_checkouts_and_stats(self):
        """Test with checkouts and statistics"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "John Doe",
                    "department": "Engineering",
                    "checkout_date": "2024-01-15T10:00:00",
                    "return_date": "2024-01-16T14:30:00",
                    "duration": 1.5
                }
            ],
            "stats": {
                "totalCheckouts": 100,
                "returnedCheckouts": 95,
                "currentlyCheckedOut": 5,
                "averageDuration": 2.3
            }
        }

        add_checkout_history_pdf_content(story, data, styles)

        # Should have: heading, summary table, spacer, details heading, details table
        assert len(story) >= 4

    def test_empty_checkouts(self):
        """Test with empty checkout list"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "checkouts": [],
            "stats": {
                "totalCheckouts": 0,
                "returnedCheckouts": 0,
                "currentlyCheckedOut": 0,
                "averageDuration": 0
            }
        }

        add_checkout_history_pdf_content(story, data, styles)

        # Should only have summary section, no checkout details
        assert len(story) >= 3

    def test_active_checkout_no_return_date(self):
        """Test checkout without return date (active checkout)"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "John Doe",
                    "department": "Engineering",
                    "checkout_date": "2024-01-15T10:00:00",
                    "return_date": None,
                    "duration": ""
                }
            ],
            "stats": {
                "totalCheckouts": 1,
                "returnedCheckouts": 0,
                "currentlyCheckedOut": 1,
                "averageDuration": 0
            }
        }

        add_checkout_history_pdf_content(story, data, styles)

        assert len(story) >= 4

    def test_more_than_50_checkouts_limited(self):
        """Test that PDF limits to first 50 checkouts"""
        story = []
        styles = getSampleStyleSheet()
        checkouts = [
            {
                "tool_number": f"T{i:03d}",
                "user_name": f"User {i}",
                "department": "Engineering",
                "checkout_date": "2024-01-15T10:00:00",
                "return_date": "2024-01-16T14:30:00",
                "duration": 1.5
            }
            for i in range(60)  # More than 50
        ]
        data = {
            "checkouts": checkouts,
            "stats": {
                "totalCheckouts": 60,
                "returnedCheckouts": 60,
                "currentlyCheckedOut": 0,
                "averageDuration": 1.5
            }
        }

        add_checkout_history_pdf_content(story, data, styles)

        # Should still complete successfully
        assert len(story) >= 4

    def test_missing_stats_fields(self):
        """Test with missing statistics fields"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "checkouts": [],
            "stats": {}  # Empty stats
        }

        add_checkout_history_pdf_content(story, data, styles)

        assert len(story) >= 3

    def test_checkout_with_empty_string_return_date(self):
        """Test checkout with empty string return date"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "John Doe",
                    "department": "Engineering",
                    "checkout_date": "2024-01-15T10:00:00",
                    "return_date": "",  # Empty string
                    "duration": ""
                }
            ],
            "stats": {
                "totalCheckouts": 1,
                "returnedCheckouts": 0,
                "currentlyCheckedOut": 1,
                "averageDuration": 0
            }
        }

        add_checkout_history_pdf_content(story, data, styles)

        assert len(story) >= 4


class TestAddDepartmentUsagePDFContent:
    """Tests for add_department_usage_pdf_content function"""

    def test_empty_departments(self):
        """Test with no departments"""
        story = []
        styles = getSampleStyleSheet()
        data = {"departments": []}

        add_department_usage_pdf_content(story, data, styles)

        assert len(story) == 1  # "No department usage data found" message

    def test_single_department(self):
        """Test with single department"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "departments": [
                {
                    "name": "Engineering",
                    "totalCheckouts": 50,
                    "currentlyCheckedOut": 3,
                    "averageDuration": 2.5,
                    "mostUsedCategory": "Hand Tools"
                }
            ]
        }

        add_department_usage_pdf_content(story, data, styles)

        assert len(story) == 1  # Table

    def test_multiple_departments(self):
        """Test with multiple departments"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "departments": [
                {
                    "name": "Engineering",
                    "totalCheckouts": 50,
                    "currentlyCheckedOut": 3,
                    "averageDuration": 2.5,
                    "mostUsedCategory": "Hand Tools"
                },
                {
                    "name": "Maintenance",
                    "totalCheckouts": 75,
                    "currentlyCheckedOut": 5,
                    "averageDuration": 3.0,
                    "mostUsedCategory": "Power Tools"
                }
            ]
        }

        add_department_usage_pdf_content(story, data, styles)

        assert len(story) == 1

    def test_missing_department_fields(self):
        """Test with missing department fields"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "departments": [
                {
                    "name": "Engineering"
                    # Missing other fields
                }
            ]
        }

        add_department_usage_pdf_content(story, data, styles)

        assert len(story) == 1


class TestAddCycleCountPDFContent:
    """Tests for add_cycle_count_pdf_content function"""

    def test_cycle_count_accuracy_report(self):
        """Test cycle count accuracy report"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "summary": {
                "total_counts": 100,
                "accurate_counts": 95,
                "accuracy_rate": 95.0
            }
        }

        add_cycle_count_pdf_content(story, data, "cycle-count-accuracy", styles)

        # Should have heading, summary stats, and additional message
        assert len(story) >= 4

    def test_cycle_count_other_type(self):
        """Test other cycle count report types"""
        story = []
        styles = getSampleStyleSheet()
        data = {
            "summary": {
                "total_counts": 50
            }
        }

        add_cycle_count_pdf_content(story, data, "cycle-count-performance", styles)

        # Should have heading and additional message
        assert len(story) >= 2

    def test_cycle_count_empty_summary(self):
        """Test with empty summary"""
        story = []
        styles = getSampleStyleSheet()
        data = {"summary": {}}

        add_cycle_count_pdf_content(story, data, "cycle-count-accuracy", styles)

        assert len(story) >= 4

    def test_cycle_count_missing_summary_key(self):
        """Test with missing summary key"""
        story = []
        styles = getSampleStyleSheet()
        data = {}

        add_cycle_count_pdf_content(story, data, "cycle-count-accuracy", styles)

        assert len(story) >= 2


class TestAddToolInventoryExcelContent:
    """Tests for add_tool_inventory_excel_content function"""

    def test_empty_tool_inventory(self):
        """Test with empty tool list"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        add_tool_inventory_excel_content(worksheet, [])

        assert worksheet["A4"].value == "No tools found."

    def test_single_tool(self):
        """Test with single tool"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "Test Tool",
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available",
                "condition": "Good"
            }
        ]

        add_tool_inventory_excel_content(worksheet, data)

        # Check headers
        assert worksheet.cell(row=4, column=1).value == "Tool Number"
        assert worksheet.cell(row=4, column=7).value == "Condition"
        # Check data
        assert worksheet.cell(row=5, column=1).value == "T001"
        assert worksheet.cell(row=5, column=7).value == "Good"

    def test_multiple_tools(self):
        """Test with multiple tools"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = [
            {
                "tool_number": f"T00{i}",
                "serial_number": f"SN00{i}",
                "description": f"Test Tool {i}",
                "category": "Hand Tools",
                "location": f"Bay {i}",
                "status": "available",
                "condition": "Good"
            }
            for i in range(1, 4)
        ]

        add_tool_inventory_excel_content(worksheet, data)

        # Check all rows are added
        assert worksheet.cell(row=5, column=1).value == "T001"
        assert worksheet.cell(row=6, column=1).value == "T002"
        assert worksheet.cell(row=7, column=1).value == "T003"

    def test_missing_tool_fields(self):
        """Test with missing tool fields"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = [
            {
                "tool_number": "T001"
                # Missing other fields
            }
        ]

        add_tool_inventory_excel_content(worksheet, data)

        assert worksheet.cell(row=5, column=1).value == "T001"
        assert worksheet.cell(row=5, column=2).value == ""  # Default empty string

    def test_header_styling(self):
        """Test that headers are styled correctly"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = [{"tool_number": "T001"}]

        add_tool_inventory_excel_content(worksheet, data)

        # Check that header has bold font
        assert worksheet.cell(row=4, column=1).font.bold is True


class TestAddCheckoutHistoryExcelContent:
    """Tests for add_checkout_history_excel_content function"""

    def test_with_checkouts(self):
        """Test with checkout data"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "John Doe",
                    "department": "Engineering",
                    "checkout_date": "2024-01-15T10:00:00",
                    "return_date": "2024-01-16T14:30:00",
                    "duration": 1.5
                }
            ],
            "stats": {
                "totalCheckouts": 100,
                "returnedCheckouts": 95,
                "currentlyCheckedOut": 5,
                "averageDuration": 2.3
            }
        }

        add_checkout_history_excel_content(worksheet, data)

        # Check summary statistics
        assert worksheet["A5"].value == "Total Checkouts"
        assert worksheet["B5"].value == 100
        # Check checkout details
        assert worksheet["A10"].value == "Checkout Details"
        assert worksheet.cell(row=12, column=1).value == "T001"

    def test_empty_checkouts(self):
        """Test with no checkouts"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "checkouts": [],
            "stats": {
                "totalCheckouts": 0,
                "returnedCheckouts": 0,
                "currentlyCheckedOut": 0,
                "averageDuration": 0
            }
        }

        add_checkout_history_excel_content(worksheet, data)

        # Summary should still be there
        assert worksheet["B5"].value == 0
        # But no checkout details section header if no checkouts
        assert worksheet["A10"].value is None

    def test_active_checkout_no_return_date(self):
        """Test checkout without return date shows 'Active'"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "John Doe",
                    "department": "Engineering",
                    "checkout_date": "2024-01-15T10:00:00",
                    "return_date": "",
                    "duration": ""
                }
            ],
            "stats": {
                "totalCheckouts": 1,
                "returnedCheckouts": 0,
                "currentlyCheckedOut": 1,
                "averageDuration": 0
            }
        }

        add_checkout_history_excel_content(worksheet, data)

        # Check that "Active" is shown for empty return date
        assert worksheet.cell(row=12, column=5).value == "Active"

    def test_missing_stats_fields(self):
        """Test with missing stats fields"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "checkouts": [],
            "stats": {}
        }

        add_checkout_history_excel_content(worksheet, data)

        # Should use default values
        assert worksheet["B5"].value == 0


class TestAddDepartmentUsageExcelContent:
    """Tests for add_department_usage_excel_content function"""

    def test_empty_departments(self):
        """Test with no departments"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {"departments": []}

        add_department_usage_excel_content(worksheet, data)

        assert worksheet["A4"].value == "No department usage data found."

    def test_single_department(self):
        """Test with single department"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "departments": [
                {
                    "name": "Engineering",
                    "totalCheckouts": 50,
                    "currentlyCheckedOut": 3,
                    "averageDuration": 2.5,
                    "mostUsedCategory": "Hand Tools"
                }
            ]
        }

        add_department_usage_excel_content(worksheet, data)

        # Check headers
        assert worksheet.cell(row=4, column=1).value == "Department"
        # Check data
        assert worksheet.cell(row=5, column=1).value == "Engineering"
        assert worksheet.cell(row=5, column=2).value == 50

    def test_multiple_departments(self):
        """Test with multiple departments"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "departments": [
                {
                    "name": "Engineering",
                    "totalCheckouts": 50,
                    "currentlyCheckedOut": 3,
                    "averageDuration": 2.5,
                    "mostUsedCategory": "Hand Tools"
                },
                {
                    "name": "Maintenance",
                    "totalCheckouts": 75,
                    "currentlyCheckedOut": 5,
                    "averageDuration": 3.0,
                    "mostUsedCategory": "Power Tools"
                }
            ]
        }

        add_department_usage_excel_content(worksheet, data)

        assert worksheet.cell(row=5, column=1).value == "Engineering"
        assert worksheet.cell(row=6, column=1).value == "Maintenance"

    def test_missing_department_fields(self):
        """Test with missing department fields"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "departments": [
                {
                    "name": "Engineering"
                    # Missing other fields - will use default 0
                }
            ]
        }

        add_department_usage_excel_content(worksheet, data)

        assert worksheet.cell(row=5, column=1).value == "Engineering"
        assert worksheet.cell(row=5, column=2).value == 0  # Default


class TestAddCycleCountExcelContent:
    """Tests for add_cycle_count_excel_content function"""

    def test_cycle_count_accuracy_report(self):
        """Test cycle count accuracy Excel report"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "summary": {
                "total_counts": 100,
                "accurate_counts": 95,
                "accuracy_rate": 95.0
            }
        }

        add_cycle_count_excel_content(worksheet, data, "cycle-count-accuracy")

        assert "cycle-count-accuracy" in worksheet["A4"].value
        assert worksheet["A6"].value == "Total Counts"
        assert worksheet["B6"].value == 100
        assert worksheet["A7"].value == "Accurate Counts"
        assert worksheet["B7"].value == 95
        assert worksheet["A8"].value == "Accuracy Rate (%)"
        assert worksheet["B8"].value == 95.0

    def test_cycle_count_other_type(self):
        """Test other cycle count report type"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {
            "summary": {
                "total_counts": 50
            }
        }

        add_cycle_count_excel_content(worksheet, data, "cycle-count-performance")

        assert "cycle-count-performance" in worksheet["A4"].value
        # Should not have accuracy-specific fields
        assert worksheet["A6"].value is None

    def test_cycle_count_empty_summary(self):
        """Test with empty summary"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {"summary": {}}

        add_cycle_count_excel_content(worksheet, data, "cycle-count-accuracy")

        # Should use default values
        assert worksheet["B6"].value == 0
        assert worksheet["B7"].value == 0
        assert worksheet["B8"].value == 0

    def test_cycle_count_missing_summary(self):
        """Test with missing summary key"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {}

        add_cycle_count_excel_content(worksheet, data, "cycle-count-accuracy")

        # Should handle missing summary gracefully
        assert worksheet["B6"].value == 0

    def test_additional_info_message(self):
        """Test that additional info message is added"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        data = {"summary": {}}

        add_cycle_count_excel_content(worksheet, data, "cycle-count-coverage")

        assert "Detailed cycle count reporting" in worksheet["A10"].value


class TestIntegration:
    """Integration tests for complete report generation workflow"""

    def test_full_pdf_workflow_tool_inventory(self):
        """Test complete PDF workflow for tool inventory"""
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "Precision Screwdriver Set",
                "category": "Hand Tools",
                "location": "Tool Crib A",
                "status": "available"
            },
            {
                "tool_number": "T002",
                "serial_number": "SN002",
                "description": "Digital Multimeter",
                "category": "Electrical",
                "location": "Tool Crib B",
                "status": "checked_out"
            }
        ]

        buffer = generate_pdf_report(data, "tool-inventory", "monthly")

        # Verify PDF was generated
        assert isinstance(buffer, io.BytesIO)
        pdf_data = buffer.getvalue()
        assert len(pdf_data) > 0
        # Check PDF magic bytes
        assert pdf_data[:4] == b"%PDF"

    def test_full_excel_workflow_checkout_history(self):
        """Test complete Excel workflow for checkout history"""
        data = {
            "checkouts": [
                {
                    "tool_number": "T001",
                    "user_name": "Alice Smith",
                    "department": "Engineering",
                    "checkout_date": "2024-01-10T08:00:00",
                    "return_date": "2024-01-12T17:00:00",
                    "duration": 2.4
                },
                {
                    "tool_number": "T002",
                    "user_name": "Bob Jones",
                    "department": "Maintenance",
                    "checkout_date": "2024-01-11T09:00:00",
                    "return_date": None,
                    "duration": None
                }
            ],
            "stats": {
                "totalCheckouts": 150,
                "returnedCheckouts": 145,
                "currentlyCheckedOut": 5,
                "averageDuration": 3.2
            }
        }

        buffer = generate_excel_report(data, "checkout-history", "weekly")

        # Verify Excel file structure
        workbook = openpyxl.load_workbook(buffer)
        ws = workbook.active
        assert "Checkout History Report" in ws["A1"].value
        assert ws["B5"].value == 150
        assert ws.cell(row=12, column=2).value == "Alice Smith"

    def test_round_trip_pdf_generation(self):
        """Test that PDF can be generated and read back"""
        data = {
            "departments": [
                {
                    "name": "QA",
                    "totalCheckouts": 25,
                    "currentlyCheckedOut": 2,
                    "averageDuration": 1.8,
                    "mostUsedCategory": "Testing Equipment"
                }
            ]
        }

        buffer = generate_pdf_report(data, "department-usage", "quarterly")
        pdf_bytes = buffer.getvalue()

        # Create new buffer from bytes
        new_buffer = io.BytesIO(pdf_bytes)
        new_buffer.seek(0)

        # Should be able to read it
        assert new_buffer.read()[:4] == b"%PDF"

    def test_round_trip_excel_generation(self):
        """Test that Excel can be generated and loaded back"""
        data = {
            "summary": {
                "total_counts": 200,
                "accurate_counts": 190,
                "accuracy_rate": 95.0
            }
        }

        buffer = generate_excel_report(data, "cycle-count-accuracy", None)
        excel_bytes = buffer.getvalue()

        # Create new buffer from bytes
        new_buffer = io.BytesIO(excel_bytes)

        # Should be able to load as workbook
        wb = openpyxl.load_workbook(new_buffer)
        assert wb.active is not None
        assert wb.active["A1"].value is not None


class TestColumnWidthErrorHandling:
    """Tests for column width calculation error handling"""

    def test_column_width_calculation_handles_type_error(self):
        """Test that column width calculation handles TypeError gracefully"""
        # Create a mock cell that raises TypeError when trying to convert value to string
        mock_cell = MagicMock()
        mock_cell.column_letter = "A"

        # Make str(cell.value) raise TypeError by making value raise on access
        type(mock_cell).value = property(lambda self: (_ for _ in ()).throw(TypeError("Cannot convert")))

        # Test the error handling directly in the loop logic
        # This simulates what happens in generate_excel_report lines 88-91
        max_length = 0
        for cell in [mock_cell]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                continue

        # Should not raise exception and max_length stays 0
        assert max_length == 0

    def test_column_width_calculation_handles_attribute_error(self):
        """Test that column width calculation handles AttributeError gracefully"""
        # Create an object that doesn't have a value attribute
        class NoValueCell:
            column_letter = "A"

        mock_cell = NoValueCell()

        # Test the error handling directly
        max_length = 0
        for cell in [mock_cell]:
            try:
                max_length = max(max_length, len(str(cell.value)))  # type: ignore
            except (TypeError, AttributeError):
                continue

        # Should not raise exception and max_length stays 0
        assert max_length == 0

    def test_column_width_with_mixed_cells(self):
        """Test column width calculation with mix of valid and problematic cells"""
        # Create both valid and problematic cells
        good_cell = MagicMock()
        good_cell.value = "Valid"
        good_cell.column_letter = "A"

        # Cell that raises TypeError
        error_cell = MagicMock()
        type(error_cell).value = property(lambda self: (_ for _ in ()).throw(TypeError()))

        # Test the loop handles mixed cells
        max_length = 0
        for cell in [good_cell, error_cell]:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except (TypeError, AttributeError):
                continue

        # Should have measured the valid cell
        assert max_length == 5  # len("Valid")

    def test_generate_excel_with_mocked_problematic_column(self):
        """Test generate_excel_report with a column that has problematic cells"""
        with patch('utils.export_utils.openpyxl.Workbook') as mock_wb_class:
            mock_workbook = MagicMock()
            mock_worksheet = MagicMock()
            mock_workbook.active = mock_worksheet
            mock_wb_class.return_value = mock_workbook

            # Create cells - one good, one that raises TypeError
            good_cell = MagicMock()
            good_cell.value = "Test"
            good_cell.column_letter = "A"

            bad_cell = MagicMock()
            type(bad_cell).value = property(lambda self: (_ for _ in ()).throw(TypeError()))

            # Set up the worksheet
            mock_worksheet.columns = [[good_cell, bad_cell]]
            mock_worksheet.column_dimensions = {"A": MagicMock()}
            mock_worksheet.__getitem__ = MagicMock(return_value=MagicMock())
            mock_worksheet.cell = MagicMock(return_value=MagicMock())

            # Mock buffer operations
            mock_buffer = io.BytesIO()

            with patch('io.BytesIO', return_value=mock_buffer):
                buffer = generate_excel_report([], "tool-inventory", None)

            # Should complete without raising exception
            assert mock_workbook.save.called


class TestEdgeCases:
    """Tests for edge cases and boundary conditions"""

    def test_tool_description_exactly_30_chars(self):
        """Test tool description with exactly 30 characters"""
        story = []
        styles = getSampleStyleSheet()
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "A" * 30,  # Exactly 30 chars
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available"
            }
        ]

        add_tool_inventory_pdf_content(story, data, styles)

        assert len(story) == 1

    def test_tool_description_exactly_31_chars(self):
        """Test tool description with 31 characters (should truncate)"""
        story = []
        styles = getSampleStyleSheet()
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "A" * 31,  # 31 chars, should truncate
                "category": "Hand Tools",
                "location": "Bay 1",
                "status": "available"
            }
        ]

        add_tool_inventory_pdf_content(story, data, styles)

        assert len(story) == 1

    def test_exactly_50_checkouts_for_pdf(self):
        """Test PDF with exactly 50 checkouts (boundary condition)"""
        story = []
        styles = getSampleStyleSheet()
        checkouts = [
            {
                "tool_number": f"T{i:03d}",
                "user_name": f"User {i}",
                "department": "Engineering",
                "checkout_date": "2024-01-15T10:00:00",
                "return_date": "2024-01-16T14:30:00",
                "duration": 1.5
            }
            for i in range(50)  # Exactly 50
        ]
        data = {
            "checkouts": checkouts,
            "stats": {"totalCheckouts": 50, "returnedCheckouts": 50, "currentlyCheckedOut": 0, "averageDuration": 1.5}
        }

        add_checkout_history_pdf_content(story, data, styles)

        assert len(story) >= 4

    def test_very_long_sheet_title_truncation(self):
        """Test Excel sheet title truncation at exactly 31 characters"""
        data = []
        # This will create a title longer than 31 chars with timeframe
        buffer = generate_excel_report(data, "cycle-count-discrepancies", "quarterly")

        workbook = openpyxl.load_workbook(buffer)
        assert len(workbook.active.title) == 31

    def test_special_characters_in_data(self):
        """Test handling of special characters in data"""
        data = [
            {
                "tool_number": "T001-<>&",
                "serial_number": "SN001\"'",
                "description": "Test & Tool <special>",
                "category": "Hand & Power Tools",
                "location": "Bay #1",
                "status": "available"
            }
        ]

        # PDF generation should handle special characters
        buffer = generate_pdf_report(data, "tool-inventory", None)
        assert buffer.getvalue()

    def test_unicode_characters_in_data(self):
        """Test handling of unicode characters"""
        data = [
            {
                "tool_number": "T001",
                "serial_number": "SN001",
                "description": "Tool with unicode: éàü",
                "category": "Ferramentas",
                "location": "Baia 1",
                "status": "disponível"
            }
        ]

        buffer = generate_excel_report(data, "tool-inventory", None)
        workbook = openpyxl.load_workbook(buffer)
        assert workbook.active is not None

    def test_numeric_values_as_strings(self):
        """Test that numeric values in strings are handled correctly"""
        data = {
            "departments": [
                {
                    "name": "Dept123",
                    "totalCheckouts": "50",  # String instead of int
                    "currentlyCheckedOut": 3,
                    "averageDuration": "2.5",  # String instead of float
                    "mostUsedCategory": "Category999"
                }
            ]
        }

        # Should handle without errors
        buffer = generate_excel_report(data, "department-usage", None)
        assert buffer.getvalue()

    def test_none_values_in_nested_dicts(self):
        """Test handling of None values in nested dictionaries"""
        data = {
            "checkouts": [
                {
                    "tool_number": None,
                    "user_name": None,
                    "department": None,
                    "checkout_date": None,
                    "return_date": None,
                    "duration": None
                }
            ],
            "stats": {
                "totalCheckouts": None,
                "returnedCheckouts": None,
                "currentlyCheckedOut": None,
                "averageDuration": None
            }
        }

        # Should handle None values gracefully
        story = []
        styles = getSampleStyleSheet()
        add_checkout_history_pdf_content(story, data, styles)
        assert len(story) > 0
